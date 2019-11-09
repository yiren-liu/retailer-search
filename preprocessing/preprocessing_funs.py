import ast
import csv
import json
from openpyxl import Workbook,load_workbook

def read_excel_to_label(excel_file,label_file):
    with open(label_file,'a',encoding='utf-8') as f:
        wb=load_workbook(excel_file)
        ws = wb.active
        for i,cell in enumerate(ws['C']):
            if cell.value!=None:
                f.write('\t'.join(ast.literal_eval(ws['A%d'%(i+1)].value))+'\t'+str(cell.value)+'\n')





def read_description_file(description_file):

    with open(description_file, 'r', encoding='utf-8') as f:
        search_results = csv.reader(f, delimiter='\t')
        all_description={}
        all_title={}
        for line in search_results:
            try:
                all_description[(line[0], line[1],line[2])]=line[4]
                all_title[(line[0], line[1],line[2])]=line[3]
            except Exception:
                print('error')
                continue
    return all_title,all_description

def creat_excel_for_labeled():
    wb = Workbook()
    ws = wb.create_sheet('sheet')
    file_total = list(csv.reader(open('../data/excel_label_data/total.csv', 'r', encoding='utf-8'),delimiter='\t'))
    file_l = open('../data/excel_label_data/file1.en.zh-CN.txt', 'r', encoding='utf-8').readlines()
    file_2 = open('../data/excel_label_data/file2.en.zh-CN.txt', 'r', encoding='utf-8').readlines()
    file_3 = open('../data/excel_label_data/file3.en.zh-CN.txt', 'r', encoding='utf-8').readlines()
    file_4 = open('../data/excel_label_data/file4.en.zh-CN.txt', 'r', encoding='utf-8').readlines()

    for i in range(len(file_total)):
        ws.append([str(tuple(file_total[i][:3])),
                   '\n'.join(file_total[i][3:6])+'\n\n'+'\n'.join(file_total[i][6:8]),
                   file_l[i]+file_total[i][4]+file_2[i]+'\n'+file_3[i]+file_4[i]])
    wb.save('label.xlsx')


def read_index_file(file):
    with open(file, 'r', encoding='utf-8') as f:
        lines=csv.reader(f, delimiter='\t')
        indexs=[]
        for line in lines:
            indexs.append(tuple(line))
    return set(indexs)
#[query, page, num, title1, website, abstract, title2, description]
def save_500_results(results):
    indexs_file=open('../data/excel_label_data/indexs_f.csv','a',encoding='utf-8')
    file_total=open('../data/excel_label_data/total.csv','w',encoding='utf-8')
    file_l = open('../data/excel_label_data/file1.txt', 'w', encoding='utf-8')
    file_2 = open('../data/excel_label_data/file2.txt', 'w', encoding='utf-8')
    file_3 = open('../data/excel_label_data/file3.txt', 'w', encoding='utf-8')
    file_4 = open('../data/excel_label_data/file4.txt', 'w', encoding='utf-8')
    for result in results:
        indexs_file.write('\t'.join(result[:3])+'\n')
        file_total.write('\t'.join(result)+'\n')
        file_l.write(result[3]+'\n')
        file_2.write(result[5]+'\n')
        file_3.write(result[6]+'\n')
        file_4.write(result[7]+'\n')
    file_total.close()
    file_l.close()
    file_2.close()
    file_3.close()
    file_4.close()

def read_label_file():
    label_file = '../data/label_data_retailer_categories.csv'
    with open(label_file, 'r', encoding='utf-8') as f:
        labels={}
        search_results = csv.reader(f, delimiter='\t')
        for line in search_results:
            # if (line[0], line[1],line[2]) in labels.keys():
            #     continue
            labels[(line[0], line[1],line[2])]=line[3]
    return labels

def read_result_file():
    result_file = '../data/result_retailer.csv'
    with open(result_file, 'r', encoding='utf-8') as f:
        f = csv.reader(f, delimiter='\t')
        all_search_results={}
        for i, line in enumerate(f):

            one_page_results = ast.literal_eval(line[2])

            for count, one_result in enumerate(one_page_results):
                all_search_results[(line[0], line[1], str(count))] = one_result
    return all_search_results