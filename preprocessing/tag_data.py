import ast
import csv
import json
import random
from openpyxl import Workbook
from preprocessing_funs import *

#将要标注的数据写入表格中
def tag_data():
    result_file='../data/result_categories_retailer.csv'
    label_file='../data/label_data_retailer_categories.csv'
    description_file = '../data/description_categories.csv'
    index_f='../data/excel_label_data/indexs_f.csv'

    all_title,all_description=read_description_file(description_file)


    #获得标签数据
    with open(label_file, 'r', encoding='utf-8') as f:
        search_results = csv.reader(f, delimiter='\t')
        labeled_results = set([(line[0], line[1],line[2]) for line in search_results])

    all_result=[]
    # file= open('label_excel.csv','w',encoding='utf-8')



    excel_file_name='../data/index_label_excel.xlsx'
    wb = Workbook()
    ws = wb.create_sheet('sheet')
    all_excel_data=[]

    unlabeled_data=[]

    #打开搜索结果文件
    with open(result_file, 'r', encoding='utf-8') as f:
        f=csv.reader(f, delimiter='\t')
        for i,line in enumerate(f):

            one_page_results=ast.literal_eval(line[2])

            #处理one page的内容
            for count,one_result in enumerate(one_page_results):
                if (line[0],line[1],str(count)) in labeled_results:
                    continue
                if one_result in all_result:
                    continue

                if (line[0],line[1],str(count)) not in all_description.keys():#543开始
                    continue


                print('\n'+line[0]+'\n'+one_result[0]+'\n'+one_result[1]+'\n'+one_result[2])

                print('标题：',all_title[(line[0],line[1],str(count))])
                print('描述：',all_description[(line[0],line[1],str(count))])

                all_excel_data.append([str((line[0],line[1],str(count))),'\n'+line[0]+'\n'+one_result[0]+'\n'+one_result[1]+'\n'+one_result[2]
                           +'\n\n'+'标题：'+all_title[(line[0],line[1],str(count))]+'\n'+'描述：'
                           +all_description[(line[0],line[1],str(count))]])

                unlabeled_data.append([line[0],line[1],str(count),one_result[0],one_result[1],one_result[2],
                                       all_title[(line[0],line[1],str(count))],all_description[(line[0],line[1],str(count))]])


                # file.write(line[0]+'\t'+one_result[0]+'\t'+one_result[1]+'\t'
                #            +one_result[2]+'\t'+all_title[(line[0],line[1],str(count))]
                #                                           +'\t'+all_description[(line[0],line[1],str(count))]+'\n')

                # label=input('1 为零售商，0为其他 2为生产商，3为无法判断或网站描述缺失或描述没有作用为，请标注：')
                # file = open(label_file, 'a', encoding='utf-8')
                # file.write(line[0]+'\t'+line[1]+'\t'+str(count)+'\t'+str(label)+'\n')
                # file.close()

            all_result.extend(one_page_results)
    # file.close()
    # random.shuffle(all_excel_data)
    # save_500_results(unlabeled_data)

    for line in all_excel_data:
        ws.append(line)
    wb.save(excel_file_name)

def gen_excel():
    wb=Workbook()
    ws=wb.create_sheet('sheet')
    for row in ws.iter_rows(1,10,1):
        row[0].value=2
    wb.save('test.xlsx')

def check_label():
    all_title, all_description=read_description_file('../data/description_categories.csv')
    labels=read_label_file()
    all_search_results=read_result_file()
    pass
    for i in labels.keys():
        if i in all_description.keys():
            if all_description[i]!=' ':
                print('\n'+str(i)+'\n'+all_search_results[i][0]+'\n'+all_search_results[i][1]+'\n'+all_search_results[i][2])
                print('标题：', all_title[i])
                print('描述：', all_description[i])
                print('label: ',labels[i])
                print('\n')


if __name__=='__main__':
    # gen_excel()
    pass
    # check_label()
    tag_data()
    # creat_excel_for_labeled()
    # read_excel_to_label('../data/labeled_excel/label_2.xlsx','../data/label_data_retailer_categories.csv')